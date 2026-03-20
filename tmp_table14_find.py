import openpyxl 
wb=openpyxl.load_workbook('Table 14 Home Health Care Expenditures.xlsx',data_only=True) 
ws=wb['Table 14'] 
for r in range(1,140): 
    if ws.cell(r,1).value==2024: 
        print('row',r,[ws.cell(r,c).value for c in range(1,10)]) 
