import openpyxl 
wb=openpyxl.load_workbook('Table 13 Other Health, Residential, and Personal Care Expenditures.xlsx',data_only=True) 
ws=wb['Table 13'] 
cols=[ws.cell(3,c).value for c in range(1,10)] 
print(cols) 
row=None 
for r in range(1,200): 
    if ws.cell(r,1).value==2024: row=r 
print('row',row) 
vals=[ws.cell(row,c).value for c in range(1,10)] 
print(vals) 
total=vals[1]*1e9 
names=['Total','Out of pocket','Health insurance','Private health insurance','Medicare','Medicaid','Other health insurance programs','Other third party payers'] 
for i,n in enumerate(names,1): 
    v=vals[i] 
    if i==1: continue 
    print(n,round(v,1),round(v/vals[1]*100,1)) 
