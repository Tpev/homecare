import openpyxl 
wb=openpyxl.load_workbook('Table 13 Other Health, Residential, and Personal Care Expenditures.xlsx',data_only=True) 
ws=wb[wb.sheetnames[0]] 
for c in range(1,22): 
    print(c,repr(ws.cell(3,c).value),repr(ws.cell(4,c).value),repr(ws.cell(5,c).value)) 
