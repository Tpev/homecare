from __future__ import annotations

import csv
import io
import json
import re
import time
from collections import OrderedDict
from datetime import date
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


OUT_DIR = Path(__file__).resolve().parent
RUN_DATE = date.today().isoformat()

NPPES_API_URL = "https://npiregistry.cms.hhs.gov/api/"
WAKE_ZIP_ARCGIS_QUERY_URL = (
    "https://maps.wakegov.com/arcgis/rest/services/Boundaries/ZipCodes/"
    "MapServer/0/query"
)
CENSUS_BATCH_GEOCODER_URL = (
    "https://geocoding.geo.census.gov/geocoder/geographies/addressbatch"
)

NPPES_LIMIT = 200
REQUEST_SLEEP_SECONDS = 0.05
LOG_PATH = OUT_DIR / "wake-county-primary-care-provider-build.log"
GEOCODER_BATCH_SIZE = 250
GEOCODER_RETRIES = 3

PRIMARY_CARE_SEARCH_TERMS = [
    "Primary Care",
    "Family",
    "Family Medicine",
    "Internal Medicine",
    "Pediatrics",
    "General Practice",
    "Geriatric Medicine",
    "Adult Health",
    "Gerontology",
    "Federally Qualified Health Center",
    "Rural Health",
    "Community Health",
]

PRIMARY_CARE_TAXONOMY_CODES = {
    # Physician primary care.
    "207Q00000X": "Family Medicine",
    "207QA0505X": "Family Medicine, Adult Medicine",
    "207QG0300X": "Family Medicine, Geriatric Medicine",
    "207R00000X": "Internal Medicine",
    "207RA0000X": "Internal Medicine, Adolescent Medicine",
    "207RG0300X": "Internal Medicine, Geriatric Medicine",
    "208000000X": "Pediatrics",
    "2080A0000X": "Pediatrics, Adolescent Medicine",
    "208D00000X": "General Practice",
    # Advanced practice clinicians commonly used in primary care.
    "363LA2200X": "Nurse Practitioner, Adult Health",
    "363LF0000X": "Nurse Practitioner, Family",
    "363LG0600X": "Nurse Practitioner, Gerontology",
    "363LP0200X": "Nurse Practitioner, Pediatrics",
    "363LP2300X": "Nurse Practitioner, Primary Care",
    # Primary care facilities.
    "261QP2300X": "Clinic/Center, Primary Care",
    "261QF0400X": "Clinic/Center, Federally Qualified Health Center (FQHC)",
    "261QR1300X": "Clinic/Center, Rural Health",
    "261QC1500X": "Clinic/Center, Community Health",
}

HEADERS = [
    "name",
    "record_type",
    "credential",
    "npi",
    "primary_taxonomy",
    "matched_primary_care_taxonomies",
    "all_taxonomies",
    "phone",
    "fax",
    "email",
    "email_source",
    "direct_endpoint",
    "address_1",
    "address_2",
    "city",
    "state",
    "zip",
    "geocoded_county",
    "geocode_match_status",
    "nppes_last_updated",
    "source_url",
    "notes",
]


def normalize_space(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def log(message: str) -> None:
    print(message, flush=True)
    with LOG_PATH.open("a", encoding="utf-8") as handle:
        handle.write(message + "\n")


def five_digit_zip(value: Any) -> str:
    digits = re.sub(r"\D", "", normalize_space(value))
    return digits[:5]


def format_phone(value: Any) -> str:
    digits = re.sub(r"\D", "", normalize_space(value))
    if len(digits) == 10:
        return f"{digits[:3]}-{digits[3:6]}-{digits[6:]}"
    return normalize_space(value)


def nppes_get(params: dict[str, Any]) -> dict[str, Any]:
    merged = {"version": "2.1", **params}
    response = requests.get(NPPES_API_URL, params=merged, timeout=30)
    response.raise_for_status()
    return response.json()


def fetch_wake_zip_entries() -> list[dict[str, str]]:
    params = {
        "where": "1=1",
        "outFields": "ZIPCODE,ZIPNUM,ZIPNAME,NAME",
        "returnGeometry": "false",
        "f": "json",
        "orderByFields": "ZIPCODE",
    }
    response = requests.get(WAKE_ZIP_ARCGIS_QUERY_URL, params=params, timeout=30)
    response.raise_for_status()
    payload = response.json()
    entries = []
    for feature in payload.get("features", []):
        attrs = feature.get("attributes", {})
        zip_text = normalize_space(attrs.get("ZIPCODE"))
        zip_code = five_digit_zip(zip_text)
        city = normalize_space(attrs.get("NAME") or attrs.get("ZIPNAME"))
        if zip_code:
            entries.append({"city": city.title(), "zip": zip_code, "label": zip_text})
    unique = OrderedDict(((entry["city"], entry["zip"]), entry) for entry in entries)
    return list(unique.values())


def paged_nppes_search(params: dict[str, Any]) -> list[dict[str, Any]]:
    results: list[dict[str, Any]] = []
    skip = 0
    while True:
        page_params = {**params, "limit": NPPES_LIMIT, "skip": skip}
        payload = nppes_get(page_params)
        page = payload.get("results") or []
        results.extend(page)
        if len(page) < NPPES_LIMIT:
            break
        skip += NPPES_LIMIT
        time.sleep(REQUEST_SLEEP_SECONDS)
    return results


def fetch_nppes_records(wake_zip_entries: list[dict[str, str]]) -> dict[str, dict[str, Any]]:
    records: dict[str, dict[str, Any]] = {}

    searches: list[tuple[str, dict[str, Any]]] = []
    for zip_code in sorted({entry["zip"] for entry in wake_zip_entries}):
        for term in PRIMARY_CARE_SEARCH_TERMS:
            searches.append(
                (
                    f"postal_code={zip_code}, taxonomy_description={term}",
                    {
                        "state": "NC",
                        "postal_code": zip_code,
                        "taxonomy_description": term,
                    },
                )
            )

    total = len(searches)
    for index, (label, params) in enumerate(searches, start=1):
        log(f"[{index}/{total}] NPPES search {label}")
        try:
            page_results = paged_nppes_search(params)
        except Exception as exc:
            log(f"  skipped after error: {exc}")
            continue
        for record in page_results:
            npi = normalize_space(record.get("number"))
            if npi and npi not in records:
                records[npi] = record
        log(f"  {len(page_results)} rows, {len(records)} unique NPIs so far")
        time.sleep(REQUEST_SLEEP_SECONDS)

    return records


def provider_name(record: dict[str, Any]) -> tuple[str, str, str]:
    basic = record.get("basic") or {}
    enumeration_type = normalize_space(record.get("enumeration_type"))
    if enumeration_type == "NPI-2":
        name = normalize_space(basic.get("organization_name"))
        return name, "Organization", ""

    parts = [
        normalize_space(basic.get("first_name")),
        normalize_space(basic.get("middle_name")),
        normalize_space(basic.get("last_name")),
    ]
    credential = normalize_space(basic.get("credential"))
    name = " ".join(part for part in parts if part)
    if credential:
        name = f"{name}, {credential}"
    return name, "Individual", credential


def taxonomy_strings(record: dict[str, Any]) -> tuple[str, str, str]:
    taxonomies = record.get("taxonomies") or []
    all_taxonomies = []
    matched = []
    primary_taxonomy = ""
    for taxonomy in taxonomies:
        code = normalize_space(taxonomy.get("code"))
        desc = normalize_space(taxonomy.get("desc"))
        text = f"{code} - {desc}" if code else desc
        if text:
            all_taxonomies.append(text)
        if taxonomy.get("primary") and text:
            primary_taxonomy = text
        if code in PRIMARY_CARE_TAXONOMY_CODES:
            matched.append(text)
    if not primary_taxonomy and all_taxonomies:
        primary_taxonomy = all_taxonomies[0]
    return primary_taxonomy, "; ".join(matched), "; ".join(all_taxonomies)


def has_primary_care_taxonomy(record: dict[str, Any]) -> bool:
    taxonomies = record.get("taxonomies") or []
    enumeration_type = normalize_space(record.get("enumeration_type"))
    if enumeration_type == "NPI-2":
        return any(
            normalize_space(taxonomy.get("code")) in PRIMARY_CARE_TAXONOMY_CODES
            for taxonomy in taxonomies
        )

    primary_taxonomies = [
        taxonomy for taxonomy in taxonomies if bool(taxonomy.get("primary"))
    ]
    if primary_taxonomies:
        return any(
            normalize_space(taxonomy.get("code")) in PRIMARY_CARE_TAXONOMY_CODES
            for taxonomy in primary_taxonomies
        )

    return any(
        normalize_space(taxonomy.get("code")) in PRIMARY_CARE_TAXONOMY_CODES
        for taxonomy in taxonomies
    )


def collect_location_addresses(record: dict[str, Any]) -> list[dict[str, Any]]:
    locations = []
    for address in record.get("addresses") or []:
        if normalize_space(address.get("address_purpose")).upper() == "LOCATION":
            locations.append(address)
    for address in record.get("practiceLocations") or []:
        locations.append(address)
    return locations


def endpoint_email_fields(record: dict[str, Any]) -> tuple[str, str, str]:
    endpoints = record.get("endpoints") or []
    endpoint_values = []
    email_values = []
    sources = []
    for endpoint in endpoints:
        value = normalize_space(endpoint.get("endpoint"))
        if not value or "@" not in value:
            continue
        endpoint_type = normalize_space(endpoint.get("endpointTypeDescription") or endpoint.get("endpointType"))
        endpoint_values.append(value)
        if endpoint_type.upper() == "DIRECT" or "direct" in value.lower():
            sources.append("NPPES Direct messaging endpoint")
        else:
            email_values.append(value)
            sources.append("NPPES endpoint")

    direct_endpoint = "; ".join(OrderedDict.fromkeys(endpoint_values))
    email = "; ".join(OrderedDict.fromkeys(email_values))
    email_source = "; ".join(OrderedDict.fromkeys(sources))
    if not email and direct_endpoint:
        email_source = "No general email in NPPES; Direct endpoint captured separately"
    if not email and not direct_endpoint:
        email_source = "No public email in NPPES"
    return email, email_source, direct_endpoint


def build_candidate_rows(
    records: dict[str, dict[str, Any]], wake_zips: set[str]
) -> list[dict[str, str]]:
    rows_by_key: OrderedDict[tuple[str, str, str, str, str], dict[str, str]] = OrderedDict()
    for record in records.values():
        basic = record.get("basic") or {}
        if normalize_space(basic.get("status")).upper() != "A":
            continue
        if not has_primary_care_taxonomy(record):
            continue

        name, record_type, credential = provider_name(record)
        if not name:
            continue
        primary_taxonomy, matched_taxonomies, all_taxonomies = taxonomy_strings(record)
        email, email_source, direct_endpoint = endpoint_email_fields(record)
        source_url = f"https://npiregistry.cms.hhs.gov/provider-view/{record.get('number')}"

        for address in collect_location_addresses(record):
            zip_code = five_digit_zip(address.get("postal_code"))
            if zip_code not in wake_zips:
                continue

            address_1 = normalize_space(address.get("address_1"))
            address_2 = normalize_space(address.get("address_2"))
            city = normalize_space(address.get("city")).title()
            state = normalize_space(address.get("state"))
            phone = format_phone(address.get("telephone_number"))
            fax = format_phone(address.get("fax_number"))
            key = (
                normalize_space(record.get("number")),
                address_1.upper(),
                address_2.upper(),
                city.upper(),
                zip_code,
            )
            rows_by_key[key] = {
                "name": name,
                "record_type": record_type,
                "credential": credential,
                "npi": normalize_space(record.get("number")),
                "primary_taxonomy": primary_taxonomy,
                "matched_primary_care_taxonomies": matched_taxonomies,
                "all_taxonomies": all_taxonomies,
                "phone": phone,
                "fax": fax,
                "email": email,
                "email_source": email_source,
                "direct_endpoint": direct_endpoint,
                "address_1": address_1,
                "address_2": address_2,
                "city": city,
                "state": state,
                "zip": zip_code,
                "geocoded_county": "",
                "geocode_match_status": "",
                "nppes_last_updated": normalize_space(basic.get("last_updated")),
                "source_url": source_url,
                "notes": "",
            }
    return list(rows_by_key.values())


def geocode_chunk(rows: list[dict[str, str]], offset: int = 0) -> None:
    if not rows:
        return

    for row in rows:
        row["geocode_match_status"] = "NO CENSUS RESULT"
        row["notes"] = "Census geocoder returned no row; retained by Wake ZIP boundary"

    batch = io.StringIO()
    writer = csv.writer(batch, lineterminator="\n")
    for index, row in enumerate(rows):
        street = " ".join(
            part for part in [row["address_1"], row["address_2"]] if normalize_space(part)
        )
        writer.writerow([index, street, row["city"], row["state"], row["zip"]])

    files = {"addressFile": ("wake_primary_care_addresses.csv", batch.getvalue())}
    data = {"benchmark": "Public_AR_Current", "vintage": "Current_Current"}

    response = None
    for attempt in range(1, GEOCODER_RETRIES + 1):
        try:
            response = requests.post(
                CENSUS_BATCH_GEOCODER_URL,
                data=data,
                files=files,
                timeout=300,
            )
            response.raise_for_status()
            break
        except Exception as exc:
            if attempt == GEOCODER_RETRIES:
                if len(rows) > 25:
                    midpoint = len(rows) // 2
                    log(
                        f"  Census geocoder failed for rows {offset + 1}-"
                        f"{offset + len(rows)}; retrying in smaller chunks"
                    )
                    geocode_chunk(rows[:midpoint], offset=offset)
                    geocode_chunk(rows[midpoint:], offset=offset + midpoint)
                    return
                for row in rows:
                    row["geocode_match_status"] = "NOT ATTEMPTED"
                    row["notes"] = (
                        "Census geocoder chunk failed; retained by Wake ZIP boundary"
                    )
                log(
                    f"  Census geocoder failed for rows {offset + 1}-"
                    f"{offset + len(rows)} after {GEOCODER_RETRIES} attempts: {exc}"
                )
                return
            time.sleep(3 * attempt)

    if response is None:
        return

    reader = csv.reader(io.StringIO(response.text))
    for geocoded in reader:
        if len(geocoded) < 10:
            continue
        try:
            index = int(geocoded[0])
        except ValueError:
            continue
        if index >= len(rows):
            continue
        match_status = normalize_space(geocoded[2])
        state_code = normalize_space(geocoded[8])
        county_code = normalize_space(geocoded[9])
        county_name = "Wake County, NC" if state_code == "37" and county_code == "183" else ""
        if match_status.upper() == "MATCH" and not county_name:
            county_name = f"State {state_code}, County {county_code}"
        rows[index]["geocode_match_status"] = match_status
        rows[index]["geocoded_county"] = county_name
        if match_status.upper() != "MATCH":
            rows[index]["notes"] = "Census geocoder did not match address; retained by Wake ZIP boundary"
        else:
            rows[index]["notes"] = ""


def geocode_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    for offset in range(0, len(rows), GEOCODER_BATCH_SIZE):
        chunk = rows[offset : offset + GEOCODER_BATCH_SIZE]
        log(
            f"  Census geocoder rows {offset + 1}-"
            f"{min(offset + GEOCODER_BATCH_SIZE, len(rows))} of {len(rows)}"
        )
        geocode_chunk(chunk, offset=offset)
        time.sleep(REQUEST_SLEEP_SECONDS)
    return rows


def filter_to_wake_or_unmatched(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    filtered = []
    for row in rows:
        county = row.get("geocoded_county", "")
        match_status = row.get("geocode_match_status", "")
        if county == "Wake County, NC":
            filtered.append(row)
        elif match_status.upper() != "MATCH":
            filtered.append(row)
    return filtered


def write_csv(rows: list[dict[str, str]], path: Path) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=HEADERS)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(rows: list[dict[str, str]], path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Wake primary care"
    worksheet.append(HEADERS)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in rows:
        worksheet.append([row.get(header, "") for header in HEADERS])

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    widths = {
        "A": 36,
        "B": 14,
        "C": 12,
        "D": 14,
        "E": 34,
        "F": 40,
        "G": 60,
        "H": 15,
        "I": 15,
        "J": 32,
        "K": 38,
        "L": 40,
        "M": 34,
        "N": 18,
        "O": 18,
        "P": 8,
        "Q": 10,
        "R": 20,
        "S": 20,
        "T": 16,
        "U": 45,
        "V": 40,
    }
    for column_index in range(1, len(HEADERS) + 1):
        letter = get_column_letter(column_index)
        worksheet.column_dimensions[letter].width = widths.get(letter, 18)

    workbook.save(path)


def write_summary(
    rows: list[dict[str, str]], wake_zip_entries: list[dict[str, str]], path: Path
) -> None:
    organizations = sum(1 for row in rows if row["record_type"] == "Organization")
    individuals = sum(1 for row in rows if row["record_type"] == "Individual")
    with_direct = sum(1 for row in rows if row["direct_endpoint"])
    with_general_email = sum(1 for row in rows if row["email"])
    unmatched = sum(1 for row in rows if row["geocode_match_status"].upper() != "MATCH")
    zip_list = ", ".join(sorted({entry["zip"] for entry in wake_zip_entries}))
    taxonomy_list = "\n".join(
        f"- `{code}` {desc}" for code, desc in sorted(PRIMARY_CARE_TAXONOMY_CODES.items())
    )
    content = f"""# Wake County Primary Care Provider List

Date prepared: {RUN_DATE}

## Files

- `wake-county-primary-care-providers-{RUN_DATE}.xlsx`
- `wake-county-primary-care-providers-{RUN_DATE}.csv`

## Summary

- Rows: {len(rows)}
- Organizations/practices/clinics: {organizations}
- Individual clinicians: {individuals}
- Rows with a general email in NPPES endpoint data: {with_general_email}
- Rows with an NPPES Direct messaging endpoint: {with_direct}
- Rows retained by Wake ZIP but not matched by Census geocoder: {unmatched}

## Source Notes

- Provider source: CMS/NPPES NPI Registry API, version 2.1.
- Wake ZIP source: Wake County GIS `Boundaries/ZipCodes` ArcGIS service.
- County check: U.S. Census batch geocoder. Rows that geocode outside Wake County are removed; rows that do not geocode are retained when their ZIP comes from the Wake County boundary dataset.
- Public office emails are not a standard NPPES field. The spreadsheet separates general endpoint emails from Direct messaging endpoints and marks missing emails as not published in NPPES.

## Wake ZIPs Used

{zip_list}

## Primary Care Taxonomy Codes Used

{taxonomy_list}
"""
    path.write_text(content, encoding="utf-8")


def main() -> None:
    LOG_PATH.write_text(f"Wake primary care provider build started {RUN_DATE}\n", encoding="utf-8")
    log("Fetching Wake County ZIPs")
    wake_zip_entries = fetch_wake_zip_entries()
    wake_zips = {entry["zip"] for entry in wake_zip_entries}
    log(f"Found {len(wake_zips)} Wake ZIPs")

    raw_path = OUT_DIR / f"wake-county-nppes-raw-{RUN_DATE}.json"
    if raw_path.exists():
        log(f"Loading existing raw NPPES cache from {raw_path}")
        records = json.loads(raw_path.read_text(encoding="utf-8"))
    else:
        log("Fetching NPPES records")
        records = fetch_nppes_records(wake_zip_entries)
        raw_path.write_text(json.dumps(records, indent=2), encoding="utf-8")
        log(f"Saved raw NPPES cache to {raw_path}")

    log("Filtering primary care candidates")
    rows = build_candidate_rows(records, wake_zips)
    log(f"Primary care candidate location rows: {len(rows)}")
    candidates_path = OUT_DIR / f"wake-county-primary-care-candidates-{RUN_DATE}.csv"
    write_csv(rows, candidates_path)
    log(f"Wrote pre-geocode candidate rows to {candidates_path}")

    log("Geocoding candidate addresses")
    geocoded_rows = geocode_rows(rows)
    wake_rows = filter_to_wake_or_unmatched(geocoded_rows)
    wake_rows.sort(key=lambda row: (row["city"], row["name"], row["address_1"], row["npi"]))
    log(f"Wake rows after geocode filter: {len(wake_rows)}")

    csv_path = OUT_DIR / f"wake-county-primary-care-providers-{RUN_DATE}.csv"
    xlsx_path = OUT_DIR / f"wake-county-primary-care-providers-{RUN_DATE}.xlsx"
    summary_path = OUT_DIR / f"wake-county-primary-care-provider-list-{RUN_DATE}.md"

    write_csv(wake_rows, csv_path)
    write_xlsx(wake_rows, xlsx_path)
    write_summary(wake_rows, wake_zip_entries, summary_path)

    log(f"Wrote {csv_path}")
    log(f"Wrote {xlsx_path}")
    log(f"Wrote {summary_path}")


if __name__ == "__main__":
    main()
